#!/usr/bin/env python3
"""
Italian Inheritance Document Scanner
Scans a Dropbox folder (recursively), extracts text from all documents,
parses structured data (heirs, assets, valuations), and generates reports.
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

# File type handlers
def read_txt(path):
    encodings = ['utf-8', 'latin-1', 'cp1252']
    for enc in encodings:
        try:
            return Path(path).read_text(encoding=enc)
        except (UnicodeDecodeError, Exception):
            continue
    return None

def read_pdf(path):
    try:
        import pdfplumber
        text = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text.append(t)
        return '\n\n'.join(text) if text else None
    except Exception as e:
        print(f"  Warning: Could not read PDF {path}: {e}")
        return None

def read_docx(path):
    try:
        from docx import Document
        doc = Document(path)
        return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        print(f"  Warning: Could not read DOCX {path}: {e}")
        return None

def read_xlsx(path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text.append(f"--- Sheet: {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else '' for c in row]
                if any(cells):
                    text.append('\t'.join(cells))
        return '\n'.join(text) if text else None
    except Exception as e:
        print(f"  Warning: Could not read XLSX {path}: {e}")
        return None

def read_csv(path):
    try:
        return Path(path).read_text(encoding='utf-8')
    except:
        try:
            return Path(path).read_text(encoding='latin-1')
        except Exception as e:
            print(f"  Warning: Could not read CSV {path}: {e}")
            return None

def read_image(path):
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(path)
        text = pytesseract.image_to_string(img, lang='ita+eng')
        return text if text.strip() else None
    except Exception as e:
        print(f"  Warning: Could not OCR image {path}: {e}")
        return None

HANDLERS = {
    '.txt': read_txt,
    '.pdf': read_pdf,
    '.docx': read_docx,
    '.doc': read_docx,
    '.xlsx': read_xlsx,
    '.xls': read_xlsx,
    '.csv': read_csv,
    '.png': read_image,
    '.jpg': read_image,
    '.jpeg': read_image,
    '.tiff': read_image,
    '.tif': read_image,
    '.bmp': read_image,
}

DROPBOX_FOLDER = "/Users/williampark/Library/CloudStorage/Dropbox/ WILLIAM/Italian Inheritance/Italian Inheritance Documents"
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extracted_data.json")


def scan_folder(folder):
    """Recursively scan folder and extract text from all supported documents."""
    documents = []
    folder = Path(folder)

    if not folder.exists():
        print(f"Error: Folder not found: {folder}")
        sys.exit(1)

    all_files = sorted(folder.rglob('*'))
    supported = [f for f in all_files if f.is_file() and f.suffix.lower() in HANDLERS and not f.name.startswith('.')]

    if not supported:
        print(f"No supported documents found in {folder}")
        return documents

    print(f"Found {len(supported)} document(s):\n")

    for filepath in supported:
        rel_path = filepath.relative_to(folder)
        ext = filepath.suffix.lower()
        handler = HANDLERS.get(ext)

        print(f"  Reading: {rel_path}")
        text = handler(str(filepath))

        if text and text.strip():
            documents.append({
                'path': str(rel_path),
                'folder': str(rel_path.parent) if str(rel_path.parent) != '.' else 'root',
                'filename': filepath.name,
                'type': ext,
                'text': text.strip(),
                'scanned_at': datetime.now().isoformat(),
                'size_bytes': filepath.stat().st_size,
            })
            print(f"    ✓ Extracted {len(text)} chars")
        else:
            print(f"    ✗ No text extracted")

    return documents


def parse_heirs(documents):
    """Parse heir information from document text."""
    heirs = []
    for doc in documents:
        text = doc['text']
        lines = text.split('\n')
        in_heirs = False
        for line in lines:
            line = line.strip()
            if 'eredi' in line.lower() or 'heirs' in line.lower():
                in_heirs = True
                continue
            if in_heirs and line and line[0].isdigit():
                # Parse lines like: 1. Giovanni (18/02/1960), coniugato, 1 figlio;
                heir = parse_heir_line(line)
                if heir:
                    heir['source_file'] = doc['path']
                    heirs.append(heir)
            elif in_heirs and ('immobili' in line.lower() or 'beni' in line.lower() or line == ''):
                if line and 'immobili' in line.lower():
                    in_heirs = False
    return heirs


def parse_heir_line(line):
    """Parse a single heir line."""
    import re
    # Remove leading number and punctuation
    line = re.sub(r'^\d+[\.\)\s]+', '', line).strip()
    if not line:
        return None

    heir = {}

    # Extract name (first word before parenthesis or comma)
    name_match = re.match(r'([A-Za-zÀ-ÿ]+)', line)
    if name_match:
        heir['name'] = name_match.group(1)

    # Extract date of birth
    dob_match = re.search(r'\((\d{2}/\d{2}/\d{4})\)', line)
    if dob_match:
        heir['date_of_birth'] = dob_match.group(1)

    # Extract marital status
    if 'coniugat' in line.lower():
        heir['marital_status'] = 'married'
        heir['marital_status_it'] = 'coniugato/a'
    elif 'stato libero' in line.lower() or 'libero' in line.lower():
        heir['marital_status'] = 'unmarried'
        heir['marital_status_it'] = 'stato libero'
    elif 'vedov' in line.lower():
        heir['marital_status'] = 'widowed'
        heir['marital_status_it'] = 'vedovo/a'

    # Extract number of children
    children_match = re.search(r'(\d+)\s+figli[oa]?', line)
    if children_match:
        heir['num_children'] = int(children_match.group(1))

    # Detect if twins (same DOB as another — handled at report level)
    heir['raw_text'] = line.rstrip(';').strip()

    return heir if heir.get('name') else None


def parse_assets(documents):
    """Parse asset/property information from document text."""
    assets = []
    for doc in documents:
        text = doc['text']
        lines = text.split('\n')
        in_assets = False
        for line in lines:
            line = line.strip()
            if any(kw in line.lower() for kw in ['immobili', 'beni', 'properties', 'assets']):
                in_assets = True
                continue
            if in_assets and line:
                assets.append({
                    'description': line.rstrip(';').strip(),
                    'source_file': doc['path'],
                    'raw_text': line,
                })
    return assets


def generate_report(data):
    """Generate a readable report from extracted data."""
    print("\n" + "=" * 60)
    print("  ITALIAN INHERITANCE — DOCUMENT REPORT")
    print("=" * 60)
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Documents scanned: {len(data['documents'])}")
    print("=" * 60)

    # Heirs
    print("\n── HEIRS (Eredi) ─────────────────────────────────────────")
    if data['heirs']:
        for i, h in enumerate(data['heirs'], 1):
            print(f"\n  {i}. {h.get('name', 'Unknown')}")
            if h.get('date_of_birth'):
                print(f"     Born: {h['date_of_birth']}")
            if h.get('marital_status'):
                print(f"     Status: {h['marital_status']} ({h.get('marital_status_it', '')})")
            if h.get('num_children') is not None:
                children_label = 'child' if h['num_children'] == 1 else 'children'
                print(f"     Children: {h['num_children']} {children_label}")
        print(f"\n  Total heirs: {len(data['heirs'])}")
        total_grandchildren = sum(h.get('num_children', 0) for h in data['heirs'])
        print(f"  Total grandchildren: {total_grandchildren}")

        # Detect twins (same DOB)
        dobs = {}
        for h in data['heirs']:
            dob = h.get('date_of_birth', '')
            if dob:
                dobs.setdefault(dob, []).append(h['name'])
        for dob, names in dobs.items():
            if len(names) > 1:
                print(f"  Note: {', '.join(names)} share DOB {dob} (twins)")
    else:
        print("  No heirs found yet.")

    # Italian succession law summary
    if data['heirs']:
        n = len(data['heirs'])
        print("\n── ITALIAN SUCCESSION LAW (Preliminary) ──────────────────")
        print(f"\n  With {n} children as heirs:")
        if n == 1:
            print("  Legittima (forced share): 1/2 of estate")
            print("  Quota disponibile (free share): 1/2 of estate")
        else:
            print(f"  Legittima (forced share): 2/3 of estate (split equally among {n} heirs)")
            print(f"  Quota disponibile (free share): 1/3 of estate")
            share_pct = round((2/3) / n * 100, 1)
            print(f"  Each heir's minimum forced share: {share_pct}% of estate")
        print("\n  Note: If a surviving spouse exists, shares differ.")
        print("  Note: This is preliminary — actual shares depend on")
        print("  wills, donations, and full family tree analysis.")

    # Assets
    print("\n── ASSETS (Immobili / Beni) ───────────────────────────────")
    if data['assets']:
        for i, a in enumerate(data['assets'], 1):
            print(f"  {i}. {a['description']}")
    else:
        print("  No assets documented yet.")
        print("  (Add property documents to the Dropbox folder)")

    # Documents inventory
    print("\n── DOCUMENTS ─────────────────────────────────────────────")
    for doc in data['documents']:
        print(f"  [{doc['type']}] {doc['path']} ({doc['size_bytes']} bytes)")
        if doc['folder'] != 'root':
            print(f"        Folder: {doc['folder']}")

    print("\n" + "=" * 60)
    print("  To update: add documents to Dropbox folder, then re-run:")
    print(f"  python3 scan.py")
    print("=" * 60 + "\n")


def save_data(data):
    """Save extracted data to JSON for Claude Code to reference."""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Data saved to {DATA_FILE}")


def main():
    folder = DROPBOX_FOLDER

    # Allow override via command line
    if len(sys.argv) > 1:
        folder = sys.argv[1]

    print(f"Scanning: {folder}\n")

    documents = scan_folder(folder)

    if not documents:
        print("\nNo documents with extractable text found.")
        print(f"Add documents to: {folder}")
        return

    heirs = parse_heirs(documents)
    assets = parse_assets(documents)

    data = {
        'scan_date': datetime.now().isoformat(),
        'source_folder': str(folder),
        'documents': documents,
        'heirs': heirs,
        'assets': assets,
    }

    save_data(data)
    generate_report(data)


if __name__ == '__main__':
    main()
